# -*- coding: utf-8 -*-
"""
Created on Wed Apr 19 21:29:46 2023

@author: Dongjae
"""

import os
import tempfile
import subprocess
from docx import Document
from openpyxl import Workbook

try:
    os.chdir("C://Users//Dongjae//Downloads")  # Change Directory
except:
    print("Directory not found")

def hwp_to_odt(hwp_path, odt_path):
    try:
        output = subprocess.check_output(['hwp5odt', hwp_path, '--output', odt_path], stderr=subprocess.STDOUT)
    except subprocess.CalledProcessError as e:
        print(f"Error: Failed to convert {hwp_path} to ODT.\nError message: {e.output.decode('utf-8')}")
        return False
    return True

def extract_hwp_text(hwp_path, odt_path, pandoc_path):
    if hwp_to_odt(hwp_path, odt_path):
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                # Extract the text from the ODT file
                text_file_path = os.path.join(temp_dir, 'output.txt')
                subprocess.check_call([pandoc_path, '-s', odt_path, '-t', 'plain', '-o', text_file_path])

                # Read the extracted text
                with open(text_file_path, 'r', encoding='utf-8') as text_file:
                    text = text_file.read()

            return text
        except subprocess.CalledProcessError:
            print(f"Error: Failed to extract text from {odt_path}.")
            return None
    else:
        return None

def find_files_with_specific_words(directory, words, pandoc_path):
    files_with_words = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith(".hwp"):
                filepath = os.path.join(root, file)
                odt_path = os.path.splitext(filepath)[0] + '.odt'
                hwp_content = extract_hwp_text(filepath, odt_path, pandoc_path)
                if hwp_content is not None and all(word in hwp_content for word in words):
                    files_with_words.append(filepath)
                os.remove(odt_path)  # Remove the ODT file after processing
    return files_with_words

search_directory = "C://Users//Dongjae//Downloads"  # Set the directory where HWP files are located
specific_words = ['기간', 'word2']  # Set the list of specific words to search for in HWP files
pandoc_path = "C://ProgramData//Anaconda3//Scripts//pandoc.exe"  # Set the path of pandoc executable

files = find_files_with_specific_words(search_directory, specific_words, pandoc_path)

def write_files_to_excel(files, output_excel_file, specific_words, pandoc_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Files"

    for index, file in enumerate(files):
        sheet_name = f'Sheet{index + 1}'
        ws.cell(row=index + 1, column=1, value=sheet_name)
        ws.cell(row=index + 1, column=2, value=file)

        odt_path = os.path.splitext(file)[0] + '.odt'
        hwp_content = extract_hwp_text(file, odt_path, pandoc_path)

        for row_index, line in enumerate(hwp_content.split('\n')):
            ws.cell(row=row_index + 1, column=1, value=line)

    wb.save(output_excel_file)

output_excel_file = 'output.xlsx'  # Set the output Excel file name

write_files_to_excel(files, output_excel_file, specific_words, pandoc_path)
# %%
import os
from docx import Document
from openpyxl import Workbook
from pyhwp import hwp5
from pyhwp.hwp5odt import section_to_text

try:
    os.chdir("C://Users//Dongjae//Downloads")  # Change Directory
except:
    print("Directory not found")

def extract_hwp_text(hwp_path):
    try:
        hwp_file = hwp5.File(hwp_path)
        sections = hwp_file.bodytext.section_streams()
        text = ""
        for section in sections:
            text += section_to_text(section)

        return text
    except Exception as e:
        print(f"Error: Failed to extract text from {hwp_path}.\nError message: {str(e)}")
        return None

def find_files_with_specific_words(directory, words):
    files_with_words = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith(".hwp"):
                filepath = os.path.join(root, file)
                hwp_content = extract_hwp_text(filepath)
                if hwp_content is not None and all(word in hwp_content for word in words):
                    files_with_words.append(filepath)
    return files_with_words

search_directory = "C://Users//Dongjae//Downloads"  # Set the directory where HWP files are located
specific_words = ['기간', 'word2']  # Set the list of specific words to search for in HWP files

files = find_files_with_specific_words(search_directory, specific_words)

def write_files_to_excel(files, output_excel_file, specific_words):
    wb = Workbook()
    ws = wb.active
    ws.title = "Files"

    for index, file in enumerate(files):
        sheet_name = f'Sheet{index + 1}'
        ws.cell(row=index + 1, column=1, value=sheet_name)
        ws.cell(row=index + 1, column=2, value=file)

        hwp_content = extract_hwp_text(file)

        for row_index, line in enumerate(hwp_content.split('\n')):
            ws.cell(row=row_index + 1, column=1, value=line)

    wb.save(output_excel_file)

output_excel_file = 'output.xlsx'  # Set the output Excel file name

write_files_to_excel(files, output_excel_file, specific_words)

# %%
import site
print(site.getsitepackages())

import sys
sys.path.append('C:\\ProgramData\\Anaconda3\\lib\\site-packages\\pyhwp')

